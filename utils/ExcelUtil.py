import openpyxl
import os
from config import Conf


class SheetTypeError:
    pass


class ExcelReader:
    def __init__(self, excel_file, sheet_by):
        if os.path.exists(excel_file):
            self.excel_file = excel_file
            self.sheet_by = sheet_by
            self.data_list = []
        else:
            raise FileNotFoundError("文件不存在")

    def data(self):
        # 打开Excel文件
        # self.excel_file = 'testdata.xlsx'  # 用你的Excel文件的实际文件名替换这里
        workbook = openpyxl.load_workbook(self.excel_file)
        if type(self.sheet_by) not in [str, int]:
            raise SheetTypeError("请输入Int or Str")
        elif type(self.sheet_by) == int:
            sheet = workbook[workbook.sheetnames[self.sheet_by]]
        elif type(self.sheet_by) == str:
            sheet = workbook[self.sheet_by]

            # 获取列标题
            columns = [cell.value for cell in sheet[1]]

            # 遍历每一行，将数据转换为字典
            # data_list = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(columns, row))
                self.data_list.append(row_data)
            return self.data_list


if __name__ == '__main__':
    reader = ExcelReader(Conf.get_excel_data_file(), "登录")
    print(reader.data())
