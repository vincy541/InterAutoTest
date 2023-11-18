# 1、导入包，xlrd
# import xlrd
#
# # 2、创建workbook对象
# book = xlrd.open_workbook("testdata.xlsx")
# # 3、sheet对象
# # 索引
# sheet = book.sheet_by_index(0)
# # 名称
# # sheet = book.sheet_by_name("登录")
# # 4、获取行数和列数
# rows = sheet.nrows  # 行数
# cols = sheet.ncols  # 列数
# # 5、读取每行的内容
# for r in range(rows):
#     r_values = sheet.row_values(r)
#     print(r_values)
# # 6、读取每列的内容
# 7、读取固定列的内容

import openpyxl

# 打开 Excel 文件
workbook = openpyxl.load_workbook('data.xlsx')

# 选择要读取的工作表
# worksheet = workbook.active  # 默认选择第一个工作表

# 读取第一个工作表
# worksheet = workbook[workbook.sheetnames[0]]

# 根据工作表名称读取
worksheet = workbook["登录"]
# 存储查询结果的列表
results = []

# 1.读取行数据
for row in worksheet.iter_rows(values_only=True):
    row_data = []
    for cell_value in row:
        row_data.append(cell_value)
    results.append(row_data)

# 打印查询结果
for row in results:
    print(row)

# 2.存储每列数据的列表，每个元素都是一列的数据
column_lists = []
# 读取列数据
for col_index in range(1, worksheet.max_column + 1):
    column_data = []
    for cell in worksheet.iter_rows(min_col=col_index, max_col=col_index, values_only=True):
        column_data.append(cell[0])
    column_lists.append(column_data)

# 打印每列数据
for index, column_data in enumerate(column_lists, start=1):
    print(f'Column {index}: {column_data}')

# 3.读取固定单元格
print(worksheet.cell(2, 2).value)

# 4.与标题行结合，遍历行数据
# 获取列标题
columns = [cell.value for cell in worksheet[1]]

# 遍历每一行，将数据转换为字典
data_list = []
for row in worksheet.iter_rows(min_row=2, values_only=True):
    row_data = dict(zip(columns, row))
    data_list.append(row_data)

print(data_list)
